import openpyxl
class LoginPageData:
    @staticmethod
    def getLoginData(test_case_name):
        Dict = {}
        workbook = openpyxl.load_workbook("../Testdata/logindata.xlsx")
        sheet = workbook.active
        rows = sheet.max_row
        columns= sheet.max_column
        for i in range (1,rows+1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range (2, columns+1):
                    Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
        return [Dict]

